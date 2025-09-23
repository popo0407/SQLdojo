#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Excelファイルの内容を分析してグラフが含まれているかチェック"""
import openpyxl
import os

def analyze_excel_file(filename):
    """Excelファイルの構造を分析"""
    print(f"\n=== {filename} の分析 ===")
    if not os.path.exists(filename):
        print(f"ファイルが見つかりません: {filename}")
        return
    
    try:
        wb = openpyxl.load_workbook(filename)
        print(f"ワークシート数: {len(wb.worksheets)}")
        
        for sheet in wb.worksheets:
            print(f"シート名: {sheet.title}")
            
            # データの行数・列数を確認
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"  データ範囲: {max_row}行 × {max_col}列")
            
            # ヘッダー行を確認
            if max_row > 0:
                header_row = [cell.value for cell in sheet[1]]
                print(f"  ヘッダー: {header_row}")
            
            # データの最初の数行を確認
            if max_row > 1:
                data_rows = []
                for row_num in range(2, min(max_row + 1, 5)):  # 最初の3行のデータ
                    row_data = [cell.value for cell in sheet[row_num]]
                    data_rows.append(row_data)
                print(f"  データ例: {data_rows}")
            
            # グラフがあるかチェック（複数の方法で試行）
            chart_count = 0
            chart_details = []
            
            # 方法1: _charts属性
            if hasattr(sheet, '_charts'):
                charts1 = sheet._charts
                chart_count += len(charts1)
                for i, chart in enumerate(charts1):
                    chart_details.append(f"_charts[{i}]: {type(chart).__name__}")
            
            # 方法2: drawing属性経由
            if hasattr(sheet, 'drawing'):
                drawing = sheet.drawing
                if hasattr(drawing, '_charts'):
                    charts2 = drawing._charts
                    chart_count += len(charts2)
                    for i, chart in enumerate(charts2):
                        chart_details.append(f"drawing._charts[{i}]: {type(chart).__name__}")
            
            # 方法3: 直接チャートリスト属性
            if hasattr(sheet, 'charts'):
                charts3 = sheet.charts
                chart_count += len(charts3)
                for i, chart in enumerate(charts3):
                    chart_details.append(f"charts[{i}]: {type(chart).__name__}")
            
            print(f"  グラフ数: {chart_count}")
            for detail in chart_details:
                print(f"    {detail}")
            
            # 他の要素もチェック
            if hasattr(sheet, '_images') and sheet._images:
                print(f"  画像数: {len(sheet._images)}")
            if hasattr(sheet, '_drawing') and sheet._drawing:
                print(f"  図形描画オブジェクト: あり")
        
        wb.close()
        
        # ファイルサイズも表示
        file_size = os.path.getsize(filename)
        print(f"ファイルサイズ: {file_size} bytes")
        
    except Exception as e:
        print(f"ファイル分析エラー: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    print("Excel ファイル分析開始")
    analyze_excel_file("test_no_chart.xlsx")
    analyze_excel_file("test_with_chart.xlsx")
    
    # サイズ比較
    if os.path.exists("test_no_chart.xlsx") and os.path.exists("test_with_chart.xlsx"):
        size_no_chart = os.path.getsize("test_no_chart.xlsx")
        size_with_chart = os.path.getsize("test_with_chart.xlsx")
        
        print(f"\n=== サイズ比較 ===")
        print(f"chart_configなし: {size_no_chart} bytes")
        print(f"chart_config付き: {size_with_chart} bytes")
        print(f"差分: {size_with_chart - size_no_chart} bytes")
        
        if size_with_chart > size_no_chart:
            print("✅ chart_config付きの方が大きい（グラフが含まれている可能性）")
        elif size_with_chart == size_no_chart:
            print("⚠️ サイズが同じ（グラフが生成されていない可能性）")
        else:
            print("❌ chart_config付きの方が小さい（予期しない結果）")